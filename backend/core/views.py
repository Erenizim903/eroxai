import uuid
import io
import json
from pathlib import Path
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.mail import send_mail
from django.core.signing import Signer, BadSignature
from django.core.files import File
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Count, Q, Sum
from datetime import timedelta
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

from .models import Document, DocumentTemplate, PremiumKey, PremiumKeyRequest, TemplateFill, UsageLog, UserActivityLog, AIChatLog, SiteSettings, TemplateField, UserProfile, SupportRequest
from .permissions import IsVerifiedEmail, IsAdminUser
from .serializers import (
    RegisterSerializer,
    UserSerializer,
    DocumentSerializer,
    DocumentTemplateSerializer,
    PremiumKeyRedeemSerializer,
    PremiumKeyRequestSerializer,
    TemplateFillSerializer,
    SiteSettingsSerializer,
    UsageLogSerializer,
    UserActivityLogSerializer,
    AIChatLogSerializer,
    TemplateFieldSerializer,
    ProfileUpdateSerializer,
    ApiKeysSerializer,
    SupportRequestSerializer,
)
from .services.file_processing import get_file_type
from .services.ocr_service import run_google_vision_ocr
from .services.openai_service import translate_text
from .services.chat_service import chat_with_ai, get_chat_provider_settings
from .services.template_render import render_pdf, render_xlsx, render_blank, render_free_text
from .services.utils import validate_email_domain
from openpyxl import load_workbook
from PIL import Image
import xlrd

signer = Signer()


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def issue_tokens(user):
    refresh = RefreshToken.for_user(user)
    return {"refresh": str(refresh), "access": str(refresh.access_token)}


def check_usage_limit(user):
    profile = user.profile
    if profile.is_premium:
        return True
    return profile.usage_count < settings.FREE_USAGE_LIMIT


def increment_usage(user, action, metadata=None):
    profile = user.profile
    profile.usage_count += 1
    profile.save(update_fields=["usage_count"])
    UsageLog.objects.create(user=user, action=action, metadata=metadata or {})


@api_view(["POST"])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    email = serializer.validated_data["email"]
    if not validate_email_domain(email):
        return Response({"error": "Email domain is invalid."}, status=status.HTTP_400_BAD_REQUEST)
    user = serializer.save()
    
    # Log registration activity
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    UserActivityLog.objects.create(
        user=user,
        action='register',
        ip_address=ip_address,
        user_agent=user_agent,
        metadata={'email': email}
    )
    
    # Yeni kullanıcıya otomatik kısıtlı key ver (1 kullanım hakkı)
    key_code = f"TRIAL-{uuid.uuid4().hex[:12].upper()}"
    trial_key = PremiumKey.objects.create(code=key_code, max_uses=1, is_active=True)
    
    token = signer.sign(str(user.id))
    verify_url = f"{request.data.get('verifyBaseUrl','')}/verify-email?token={token}"
    if settings.EMAIL_HOST:
        send_mail(
            "Verify your account",
            f"Click to verify: {verify_url}\n\nYour trial premium key: {key_code}\nPlease save this key and enter it in the dashboard after login.",
            settings.DEFAULT_FROM_EMAIL,
            [email],
            fail_silently=True,
        )
    return Response({
        "message": "Registered. Verify your email.",
        "token": token,
        "trial_key": key_code,
        "note": "Please save your trial premium key and enter it in the dashboard after login."
    })


@api_view(["POST"])
@permission_classes([AllowAny])
def verify_email(request):
    token = request.data.get("token")
    if not token:
        return Response({"error": "Token required."}, status=status.HTTP_400_BAD_REQUEST)
    try:
        user_id = signer.unsign(token)
    except BadSignature:
        return Response({"error": "Invalid token."}, status=status.HTTP_400_BAD_REQUEST)
    user = get_object_or_404(User, id=user_id)
    user.profile.is_email_verified = True
    user.profile.save(update_fields=["is_email_verified"])
    return Response({"message": "Email verified."})


@api_view(["POST"])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")
    user = authenticate(username=username, password=password)
    if not user:
        ip_address = get_client_ip(request)
        cache_key = f"login_fail:{ip_address}:{username}"
        attempts = cache.get(cache_key, 0) + 1
        cache.set(cache_key, attempts, timeout=900)
        if attempts >= 5:
            support_user = User.objects.filter(username=username).first()
            SupportRequest.objects.create(
                user=support_user,
                username=username or "",
                email=support_user.email if support_user else "",
                reason="5 başarısız giriş denemesi sonrası otomatik talep.",
                status="pending",
                ip_address=ip_address,
                user_agent=request.META.get("HTTP_USER_AGENT", ""),
            )
            return Response({"error": "Support request required.", "code": "support_required"}, status=status.HTTP_429_TOO_MANY_REQUESTS)
        return Response({"error": "Invalid credentials."}, status=status.HTTP_401_UNAUTHORIZED)
    cache.delete(f"login_fail:{get_client_ip(request)}:{username}")
    
    # Log login activity
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    UserActivityLog.objects.create(
        user=user,
        action='login',
        ip_address=ip_address,
        user_agent=user_agent,
        metadata={'username': username}
    )
    
    return Response({"user": UserSerializer(user).data, "tokens": issue_tokens(user)})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def me(request):
    return Response(UserSerializer(request.user).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def logout(request):
    """Logout endpoint - logs user logout activity"""
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    UserActivityLog.objects.create(
        user=request.user,
        action='logout',
        ip_address=ip_address,
        user_agent=user_agent,
    )
    return Response({"message": "Logged out successfully."})


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_profile(request):
    serializer = ProfileUpdateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = request.user
    profile = user.profile

    for field in ["first_name", "last_name", "email"]:
        if field in serializer.validated_data:
            setattr(user, field, serializer.validated_data[field])
    user.save()

    for field in ["phone", "company", "title", "address", "locale"]:
        if field in serializer.validated_data:
            setattr(profile, field, serializer.validated_data[field])

    if "avatar" in request.FILES:
        profile.avatar = request.FILES["avatar"]

    profile.save()
    return Response(UserSerializer(user).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def redeem_key(request):
    serializer = PremiumKeyRedeemSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    code = serializer.validated_data["code"]
    premium_key = get_object_or_404(PremiumKey, code=code, is_active=True)
    if premium_key.used_count >= premium_key.max_uses:
        return Response({"error": "Key already used."}, status=status.HTTP_400_BAD_REQUEST)
    premium_key.used_count += 1
    premium_key.save(update_fields=["used_count"])
    profile = request.user.profile
    profile.is_premium = True
    profile.save(update_fields=["is_premium"])
    return Response({"message": "Premium activated."})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def upload_document(request):
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    file_obj = request.FILES.get("file")
    if not file_obj:
        return Response({"error": "File required."}, status=status.HTTP_400_BAD_REQUEST)
    file_type = get_file_type(file_obj.name)
    if file_type == "unknown":
        return Response({"error": "Unsupported file type."}, status=status.HTTP_400_BAD_REQUEST)
    doc = Document.objects.create(
        uploaded_by=request.user,
        file=file_obj,
        file_type=file_type,
        source_language=request.data.get("source_language", "auto"),
        target_language="ja",
    )
    return Response(DocumentSerializer(doc).data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def run_ocr(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id, uploaded_by=request.user)
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    file_path = doc.file.path
    text = ""
    if doc.file_type in ["pdf", "image"]:
        text = run_google_vision_ocr(file_path, doc.file_type, request.data.get("language_hint"))
    elif doc.file_type == "xlsx":
        chunks = []
        try:
            workbook = load_workbook(file_path, data_only=True)
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    if row_text.strip():
                        chunks.append(row_text)
        except Exception:
            legacy = xlrd.open_workbook(file_path)
            for sheet in legacy.sheets():
                for row_idx in range(sheet.nrows):
                    row_values = sheet.row_values(row_idx)
                    row_text = " ".join([str(cell) for cell in row_values if cell not in ["", None]])
                    if row_text.strip():
                        chunks.append(row_text)
        text = "\n".join(chunks)
    else:
        return Response({"error": "Unsupported file type."}, status=status.HTTP_400_BAD_REQUEST)
    doc.extracted_text = text
    doc.status = "ocr_done"
    doc.save(update_fields=["extracted_text", "status"])
    increment_usage(request.user, "ocr", {"document": doc.id})
    return Response(DocumentSerializer(doc).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def translate_document(request, doc_id):
    doc = get_object_or_404(Document, id=doc_id, uploaded_by=request.user)
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    text = doc.extracted_text or request.data.get("text", "")
    translated = translate_text(text, doc.source_language, doc.target_language or "ja")
    doc.translated_text = translated
    doc.status = "translated"
    doc.save(update_fields=["translated_text", "status"])
    increment_usage(request.user, "translate", {"document": doc.id})
    return Response(DocumentSerializer(doc).data)


@api_view(["GET"])
@permission_classes([AllowAny])
def list_templates(request):
    if not DocumentTemplate.objects.exists():
        sample_dir = Path(settings.BASE_DIR) / "sample_templates"
        if sample_dir.exists():
            existing_slugs = set(DocumentTemplate.objects.values_list("slug", flat=True))
            for file_path in sample_dir.iterdir():
                if not file_path.is_file():
                    continue
                ext = file_path.suffix.lower().lstrip(".")
                if ext not in ["xlsx", "pdf"]:
                    continue
                base_name = file_path.stem
                slug_base = slugify(base_name) or f"template-{uuid.uuid4().hex[:8]}"
                if slug_base in existing_slugs:
                    continue
                slug = slug_base
                suffix = 1
                while slug in existing_slugs:
                    suffix += 1
                    slug = f"{slug_base}-{suffix}"
                with open(file_path, "rb") as handle:
                    template = DocumentTemplate(
                        name_tr=base_name,
                        name_en=base_name,
                        name_ja=base_name,
                        slug=slug,
                        template_type=ext,
                        is_active=True,
                    )
                    template.file.save(file_path.name, File(handle), save=True)
                existing_slugs.add(slug)
    templates = DocumentTemplate.objects.all().order_by("-created_at")
    return Response(DocumentTemplateSerializer(templates, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail, IsAdminUser])
def create_template(request):
    name = request.data.get("name") or request.data.get("name_tr")
    if not name:
        return Response({"error": "Template name is required."}, status=status.HTTP_400_BAD_REQUEST)

    description = request.data.get("description") or request.data.get("description_tr") or ""
    template_type = (request.data.get("template_type") or "xlsx").lower()
    output_language = request.data.get("output_language") or "ja"
    raw_schema = request.data.get("fields_schema", "{}")
    if isinstance(raw_schema, str):
        try:
            fields_schema = json.loads(raw_schema) if raw_schema else {}
        except json.JSONDecodeError:
            return Response({"error": "Invalid fields_schema JSON."}, status=status.HTTP_400_BAD_REQUEST)
    else:
        fields_schema = raw_schema or {}

    upload = request.FILES.get("file")
    image_types = {"png", "jpg", "jpeg"}
    if template_type in image_types:
        template_type = "pdf"

    template = DocumentTemplate.objects.create(
        name_tr=name,
        description_tr=description,
        template_type=template_type if template_type in {"xlsx", "pdf", "blank"} else "xlsx",
        output_language=output_language,
        fields_schema=fields_schema,
        created_by=request.user,
    )

    if upload:
        ext = Path(upload.name).suffix.lower().lstrip(".")
        if ext in image_types:
            image = Image.open(upload)
            if image.mode not in ("RGB", "L"):
                image = image.convert("RGB")
            buffer = io.BytesIO()
            image.save(buffer, format="PDF")
            buffer.seek(0)
            filename = f"{Path(upload.name).stem}.pdf"
            template.file.save(filename, ContentFile(buffer.read()), save=True)
            template.template_type = "pdf"
            template.save(update_fields=["file", "template_type"])
        else:
            template.file.save(upload.name, upload, save=True)

    return Response(DocumentTemplateSerializer(template).data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def fill_template(request, template_id):
    template = get_object_or_404(DocumentTemplate, id=template_id)
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    input_data = request.data.get("input_data", {})
    output_data = {}
    for key, value in input_data.items():
        output_data[key] = translate_text(str(value), "auto", template.output_language or "ja")

    fields = list(template.fields.all())
    if template.template_type == "pdf" and template.file:
        output_path = render_pdf(template.file.path, fields, output_data)
    elif template.template_type == "xlsx":
        output_path = render_xlsx(template.file.path if template.file else None, fields, output_data)
    else:
        output_path = render_blank(output_data, fields)

    fill = TemplateFill.objects.create(
        template=template,
        user=request.user,
        input_data=input_data,
        output_data=output_data,
    )
    with open(output_path, "rb") as handle:
        fill.output_file.save(output_path.name, File(handle), save=True)
    increment_usage(request.user, "fill_template", {"template": template.id})
    return Response(TemplateFillSerializer(fill).data, status=status.HTTP_201_CREATED)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def full_translate(request):
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    text = request.data.get("text", "")
    source = request.data.get("source_language", "auto")
    translated = translate_text(text, source, "ja")
    increment_usage(request.user, "full_translate")
    return Response({"translated_text": translated})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def blank_document(request):
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    text = request.data.get("text", "")
    source = request.data.get("source_language", "auto")
    title = request.data.get("title", "Translated Document")
    translated = translate_text(text, source, "ja")
    output_path = render_free_text(translated, title=title)
    increment_usage(request.user, "blank_document")
    return Response({"output_file": f"{settings.MEDIA_URL}outputs/{output_path.name}"})


@api_view(["GET"])
@permission_classes([AllowAny])
def site_settings(request):
    settings_obj, _ = SiteSettings.objects.get_or_create(id=1)
    return Response(SiteSettingsSerializer(settings_obj).data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsAdminUser])
def update_site_settings(request):
    settings_obj, _ = SiteSettings.objects.get_or_create(id=1)
    serializer = SiteSettingsSerializer(settings_obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(["GET", "PUT"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_api_keys(request):
    settings_obj, _ = SiteSettings.objects.get_or_create(id=1)
    if request.method == "GET":
        return Response(ApiKeysSerializer(settings_obj).data)
    serializer = ApiKeysSerializer(settings_obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(ApiKeysSerializer(settings_obj).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_ocr_status(request):
    try:
        import pytesseract
        tesseract_available = True
    except Exception:
        tesseract_available = False
    settings_obj, _ = SiteSettings.objects.get_or_create(id=1)
    google_key = settings_obj.google_vision_api_key or settings.GOOGLE_VISION_API_KEY
    return Response({
        "tesseract_available": tesseract_available,
        "google_vision_configured": bool(google_key),
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_support_requests(request):
    items = SupportRequest.objects.all().order_by("-created_at")[:500]
    return Response(SupportRequestSerializer(items, many=True).data)


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_support_request_update(request, request_id):
    item = get_object_or_404(SupportRequest, id=request_id)
    serializer = SupportRequestSerializer(item, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_users(request):
    users = User.objects.all().order_by("-date_joined")
    return Response(UserSerializer(users, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_create_user(request):
    serializer = RegisterSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    user = serializer.save()
    if request.data.get("is_staff"):
        user.is_staff = True
    if request.data.get("is_superuser"):
        user.is_superuser = True
    user.is_active = request.data.get("is_active", True)
    user.save()
    return Response(UserSerializer(user).data, status=status.HTTP_201_CREATED)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    return Response({"message": "User deleted."})


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_logs(request):
    logs = UsageLog.objects.all().order_by("-created_at")[:500]
    return Response(UsageLogSerializer(logs, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_activity_logs(request):
    """Get user activity logs (login, logout, register, etc.)"""
    user_id = request.query_params.get("user_id")
    action = request.query_params.get("action")
    limit = int(request.query_params.get("limit", 500))
    
    logs = UserActivityLog.objects.all()
    if user_id:
        logs = logs.filter(user_id=user_id)
    if action:
        logs = logs.filter(action=action)
    
    logs = logs.order_by("-created_at")[:limit]
    return Response(UserActivityLogSerializer(logs, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_ai_chat_logs(request):
    """Get AI chat logs - son 3 gün içindekiler (daha eski olanlar otomatik silindi)"""
    user_id = request.query_params.get("user_id")
    limit = int(request.query_params.get("limit", 500))
    
    logs = AIChatLog.objects.all()
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    logs = logs.order_by("-created_at")[:limit]
    return Response(AIChatLogSerializer(logs, many=True).data)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_premium_keys(request):
    if request.method == "GET":
        keys = PremiumKey.objects.all().order_by("-id")
        return Response(
            [{"id": k.id, "code": k.code, "is_active": k.is_active, "used_count": k.used_count, "max_uses": k.max_uses} for k in keys]
        )
    code = request.data.get("code")
    max_uses = int(request.data.get("max_uses", 1))
    key = PremiumKey.objects.create(code=code or uuid.uuid4().hex, max_uses=max_uses)
    return Response({"id": key.id, "code": key.code})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_add_template_field(request, template_id):
    template = get_object_or_404(DocumentTemplate, id=template_id)
    serializer = TemplateFieldSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    field = serializer.save(template=template)
    return Response(TemplateFieldSerializer(field).data, status=status.HTTP_201_CREATED)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_delete_template(request, template_id):
    template = get_object_or_404(DocumentTemplate, id=template_id)
    template.delete()
    return Response({"message": "Template deleted."})


@api_view(["DELETE"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_delete_template_field(request, template_id, field_id):
    field = get_object_or_404(TemplateField, id=field_id, template_id=template_id)
    field.delete()
    return Response({"message": "Field deleted."})


@api_view(["PUT"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_update_template_field(request, template_id, field_id):
    field = get_object_or_404(TemplateField, id=field_id, template_id=template_id)
    serializer = TemplateFieldSerializer(field, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_analytics(request):
    now = timezone.now()
    last_7_days = now - timedelta(days=7)
    last_30_days = now - timedelta(days=30)
    
    total_users = User.objects.count()
    active_users = User.objects.filter(last_login__gte=last_30_days).count()
    premium_users = UserProfile.objects.filter(is_premium=True).count()
    total_documents = Document.objects.count()
    total_templates = DocumentTemplate.objects.count()
    total_keys = PremiumKey.objects.count()
    active_keys = PremiumKey.objects.filter(is_active=True).count()
    
    usage_last_7 = UsageLog.objects.filter(created_at__gte=last_7_days).count()
    usage_last_30 = UsageLog.objects.filter(created_at__gte=last_30_days).count()
    
    # Activity logs stats
    login_count_7 = UserActivityLog.objects.filter(action='login', created_at__gte=last_7_days).count()
    login_count_30 = UserActivityLog.objects.filter(action='login', created_at__gte=last_30_days).count()
    logout_count_7 = UserActivityLog.objects.filter(action='logout', created_at__gte=last_7_days).count()
    logout_count_30 = UserActivityLog.objects.filter(action='logout', created_at__gte=last_30_days).count()
    register_count_7 = UserActivityLog.objects.filter(action='register', created_at__gte=last_7_days).count()
    register_count_30 = UserActivityLog.objects.filter(action='register', created_at__gte=last_30_days).count()
    
    # Unique IPs
    unique_ips_7 = UserActivityLog.objects.filter(created_at__gte=last_7_days).values('ip_address').distinct().count()
    unique_ips_30 = UserActivityLog.objects.filter(created_at__gte=last_30_days).values('ip_address').distinct().count()
    
    action_stats = UsageLog.objects.filter(created_at__gte=last_30_days).values('action').annotate(count=Count('id')).order_by('-count')[:10]
    
    user_registrations = User.objects.filter(date_joined__gte=last_30_days).extra(
        select={'day': 'date(date_joined)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    usage_timeline = UsageLog.objects.filter(created_at__gte=last_7_days).extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(count=Count('id')).order_by('day')
    
    return Response({
        "overview": {
            "total_users": total_users,
            "active_users": active_users,
            "premium_users": premium_users,
            "total_documents": total_documents,
            "total_templates": total_templates,
            "total_keys": total_keys,
            "active_keys": active_keys,
            "usage_last_7_days": usage_last_7,
            "usage_last_30_days": usage_last_30,
            "login_count_7_days": login_count_7,
            "login_count_30_days": login_count_30,
            "logout_count_7_days": logout_count_7,
            "logout_count_30_days": logout_count_30,
            "register_count_7_days": register_count_7,
            "register_count_30_days": register_count_30,
            "unique_ips_7_days": unique_ips_7,
            "unique_ips_30_days": unique_ips_30,
        },
        "action_stats": list(action_stats),
        "user_registrations": list(user_registrations),
        "usage_timeline": list(usage_timeline),
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def ai_chat(request):
    if not check_usage_limit(request.user):
        return Response({"error": "Free usage limit reached."}, status=status.HTTP_403_FORBIDDEN)
    message = request.data.get("message", "")
    language = request.data.get("language", "tr")
    provider = request.data.get("provider", None)
    if not message:
        return Response({"error": "Message required."}, status=status.HTTP_400_BAD_REQUEST)
    
    response_text = chat_with_ai(message, language, provider)
    
    # AI chat mesajını logla
    ip_address = get_client_ip(request)
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    provider_settings = get_chat_provider_settings()
    active_provider = provider or provider_settings.get('provider', 'openai')
    
    AIChatLog.objects.create(
        user=request.user,
        message=message,
        response=response_text,
        language=language,
        provider=active_provider,
        ip_address=ip_address,
        user_agent=user_agent,
    )
    
    # 3 günden eski AI chat loglarını temizle
    three_days_ago = timezone.now() - timedelta(days=3)
    AIChatLog.objects.filter(created_at__lt=three_days_ago).delete()
    
    increment_usage(request.user, "ai_chat", {"language": language})
    return Response({"response": response_text})


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def create_premium_request(request):
    """Kullanıcı premium key başvurusu yapar"""
    reason = request.data.get("reason", "")
    existing_pending = PremiumKeyRequest.objects.filter(user=request.user, status="pending").exists()
    if existing_pending:
        return Response({"error": "Zaten bekleyen bir başvurunuz var."}, status=status.HTTP_400_BAD_REQUEST)
    
    req = PremiumKeyRequest.objects.create(user=request.user, reason=reason)
    return Response(PremiumKeyRequestSerializer(req).data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsVerifiedEmail])
def my_premium_requests(request):
    """Kullanıcı kendi başvurularını görür"""
    requests = PremiumKeyRequest.objects.filter(user=request.user).order_by("-created_at")
    return Response(PremiumKeyRequestSerializer(requests, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_premium_requests(request):
    """Admin tüm başvuruları görür"""
    status_filter = request.query_params.get("status", None)
    qs = PremiumKeyRequest.objects.all().order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)
    return Response(PremiumKeyRequestSerializer(qs, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_approve_request(request, request_id):
    """Admin başvuruyu onaylar ve premium key oluşturur"""
    req = get_object_or_404(PremiumKeyRequest, id=request_id)
    if req.status != "pending":
        return Response({"error": "Bu başvuru zaten işleme alınmış."}, status=status.HTTP_400_BAD_REQUEST)
    
    # Premium key oluştur
    key_code = request.data.get("key_code") or uuid.uuid4().hex[:16]
    max_uses = int(request.data.get("max_uses", 1))
    premium_key = PremiumKey.objects.create(code=key_code, max_uses=max_uses)
    
    # Başvuruyu onayla
    req.status = "approved"
    req.approved_by = request.user
    req.admin_note = request.data.get("admin_note", "")
    req.save()
    
    # Kullanıcıya premium ver
    profile = req.user.profile
    profile.is_premium = True
    profile.save()
    
    return Response({
        "request": PremiumKeyRequestSerializer(req).data,
        "key": {"id": premium_key.id, "code": premium_key.code},
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_reject_request(request, request_id):
    """Admin başvuruyu reddeder"""
    req = get_object_or_404(PremiumKeyRequest, id=request_id)
    if req.status != "pending":
        return Response({"error": "Bu başvuru zaten işleme alınmış."}, status=status.HTTP_400_BAD_REQUEST)
    
    req.status = "rejected"
    req.approved_by = request.user
    req.admin_note = request.data.get("admin_note", "")
    req.save()
    
    return Response(PremiumKeyRequestSerializer(req).data)
