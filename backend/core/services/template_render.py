import io
import uuid
from pathlib import Path
from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook, Workbook
import xlrd


def ensure_output_dir():
    output_dir = Path(settings.MEDIA_ROOT) / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _draw_text_overlay(page, fields, output_data, page_number):
    packet = io.BytesIO()
    page_width = float(page.mediabox.width)
    page_height = float(page.mediabox.height)
    overlay = canvas.Canvas(packet, pagesize=(page_width, page_height))

    for field in fields:
        value = output_data.get(field.key)
        if value is None:
            continue
        mapping = field.mapping or {}
        if mapping.get("page", 1) != page_number:
            continue
        x = float(mapping.get("x", 72))
        y = float(mapping.get("y", 72))
        font_size = float(mapping.get("font_size", 11))
        overlay.setFont(mapping.get("font", "Helvetica"), font_size)
        overlay.drawString(x, page_height - y, str(value))

    overlay.save()
    packet.seek(0)
    return PdfReader(packet).pages[0]


def render_pdf(template_path, fields, output_data):
    reader = PdfReader(template_path)
    writer = PdfWriter()

    for index, page in enumerate(reader.pages):
        overlay_page = _draw_text_overlay(page, fields, output_data, index + 1)
        page.merge_page(overlay_page)
        writer.add_page(page)

    output_dir = ensure_output_dir()
    output_path = output_dir / f"filled-{uuid.uuid4().hex}.pdf"
    with open(output_path, "wb") as handle:
        writer.write(handle)
    return output_path


def render_blank(output_data, fields):
    output_dir = ensure_output_dir()
    output_path = output_dir / f"blank-{uuid.uuid4().hex}.pdf"
    c = canvas.Canvas(str(output_path), pagesize=A4)
    y = 800
    for field in fields:
        value = output_data.get(field.key)
        if value:
            c.drawString(72, y, f"{field.label}: {value}")
            y -= 20
    c.save()
    return output_path


def render_free_text(text, title=None):
    output_dir = ensure_output_dir()
    output_path = output_dir / f"free-{uuid.uuid4().hex}.pdf"
    c = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    y = height - 30 * mm
    if title:
        c.setFont("Helvetica-Bold", 14)
        c.drawString(20 * mm, y, title)
        y -= 12 * mm
    c.setFont("Helvetica", 11)
    max_width = width - 40 * mm
    line = ""
    for word in text.split():
        test = f"{line} {word}".strip()
        if c.stringWidth(test, "Helvetica", 11) <= max_width:
            line = test
        else:
            c.drawString(20 * mm, y, line)
            y -= 7 * mm
            if y < 20 * mm:
                c.showPage()
                y = height - 30 * mm
            line = word
    if line:
        c.drawString(20 * mm, y, line)
    c.save()
    return output_path


def render_xlsx(template_path, fields, output_data):
    workbook = None
    if template_path:
        try:
            workbook = load_workbook(template_path)
        except Exception:
            workbook = None

    if workbook is None:
        workbook = Workbook()
        if template_path and str(template_path).lower().endswith(".xls"):
            try:
                legacy = xlrd.open_workbook(template_path)
                for sheet_name in legacy.sheet_names():
                    if sheet_name not in workbook.sheetnames:
                        workbook.create_sheet(title=sheet_name)
            except Exception:
                pass
    for field in fields:
        value = output_data.get(field.key)
        if value is None:
            continue
        mapping = field.mapping or {}
        sheet_name = mapping.get("sheet", workbook.sheetnames[0])
        cell = mapping.get("cell")
        if not cell:
            continue
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        sheet[cell] = str(value)
    output_dir = ensure_output_dir()
    output_path = output_dir / f"filled-{uuid.uuid4().hex}.xlsx"
    workbook.save(output_path)
    return output_path
